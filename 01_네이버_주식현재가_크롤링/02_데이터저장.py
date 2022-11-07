# 실행전 주의사항
# fpath의 data.xlsx 위치를 본인에 맞게 수정할 것

import requests
from bs4 import BeautifulSoup
import openpyxl
import os

# fpath = r'C:\Users\hosoo.kang\dev\crawling\01_네이버_주식현재가_크롤링\data.xlsx'
fpath = r'K:\My files\PA개인폴더(stand.lee)\88. AWS Toy Project\Crawling\data_standlee.xlsx'

# 파일 있음 load, 없음 생성
if os.path.isfile(fpath) :
    wb = openpyxl.load_workbook(fpath)
else :  
    wb = openpyxl.Workbook()
# wb = openpyxl.load_workbook()
ws = wb.active  # 현재 활성화된 시트 선택

# 종목코드리스트
codes = [
    '005930', #삼성전자
    '000660', #하이닉스
    '035720'  #카카오
]

# 타이틀
ws['A1'] = '종목코드'
ws['B1'] = '종목명'
ws['C1'] = '현재가'

row = 2
for code in codes:
    url_name = f"https://finance.naver.com/item/fchart.naver?code={code}"
    response_name = requests.get(url_name)
    html_name = response_name.text
    soup_name = BeautifulSoup(html_name, 'html.parser')
    name = soup_name.select_one("h2>a").text # class 값은 앞에 . 붙여줌
    print(name)

    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    price = soup.select_one("#_nowVal").text  # id 값은 앞에 # 붙여줌
    price = price.replace(',', '')  # , 제거
    print(code, name, price, sep=", ", end="\n")
    ws[f'A{row}'] = code
    ws[f'B{row}'] = name
    ws[f'C{row}'] = int(price)
    row = row + 1

wb.save(fpath)