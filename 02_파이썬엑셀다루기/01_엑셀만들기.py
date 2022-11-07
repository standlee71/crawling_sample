import openpyxl

# 1)엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크쉬트 만들기
ws = wb.create_sheet('오징어게임')

# 3) 데이터 추가하기
ws['A1'] = '참가번호'
ws['B1'] = '성명'

ws['A2'] = 1
ws['B2'] = '오일남'

# 4) 엑셀저장하기
wb.save(r'K:\My files\PA개인폴더(stand.lee)\88. AWS Toy Project\Crawling\참가자_standlee.xlsx')  # r은 raw의 의미로 경로의 `\` 인식을 위해 붙임
