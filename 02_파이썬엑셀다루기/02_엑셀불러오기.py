# 실행전 주의사항
# fpath의 '참가자_data.xlsx' 위치를 본인에 맞게 수정할 것
import openpyxl

fpath = r'K:\My files\PA개인폴더(stand.lee)\88. AWS Toy Project\Crawling\참가자_standlee.xlsx'
# 1) 엑셀불러오기
wb = openpyxl.load_workbook(fpath)

# 2) 엑셀 시트선택
ws = wb['오징어게임']

# 3) 데이터 수정
ws['A3'] = 456
ws['B3'] = '성기훈'

# 4) 엑설저장하기
wb.save(fpath)
